import os
import time
import json
import logging
import argparse
import requests

from pathlib import Path
from openpyxl.styles import Font
from urllib.parse import urlparse
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Optional, Union

from proxy.nekoray import Proxy
from browser import MobileBrowser, Browser
from managers import DirManager, FileManager
from models import ProxyManager, ProxyUnit, ResultWorkUnit, WorkUnit, WaitWorkUnit


class WebsiteProcessor:
    def __init__(self, excel_path: str, proxy_path: str = "proxy.json") -> None:
        self._setup_logging()
        self.logger = logging.getLogger(name="processor")

        self.excel_path = Path(excel_path)
        self.proxy_path = Path(proxy_path)

        self.logger.info(f"Инициализация WebsiteProcessor с excel_path: {excel_path}, proxy_path: {proxy_path}")

        try:
            self.proxy_manager: ProxyManager = self._load_proxies()
            self.wb = load_workbook(self.excel_path)
            self.sheet = self.wb.active

            self.main_queue: List[WorkUnit] = []
            self.side_queue: List[WaitWorkUnit] = []
            self.data: List[ResultWorkUnit] = []

            self._load_main_queue()
            self.logger.info(f"Инициализация завершена успешно. Загружено {len(self.main_queue)} задач.")

        except Exception as e:
            self.logger.error(f"Ошибка инициализации: {str(e)}")
            raise

    def _setup_logging(self) -> None:
        """Настройка системы логирования."""
        logging.basicConfig(
            level=logging.INFO,
            format="[%(name)s] [%(levelname)s] - %(message)s"
        )

    def _load_proxies(self) -> ProxyManager:
        """Загрузка JSON-файла с прокси."""
        self.logger.info(f"Загрузка прокси из {self.proxy_path}")

        try:
            with self.proxy_path.open("r") as file:
                proxies = json.load(file)

            proxy_manager = ProxyManager()
            for country_name in proxies.keys():
                proxy_manager.add_proxy(
                    country_name=country_name,
                    value=ProxyUnit(**proxies[country_name])
                )

            self.logger.info(f"Успешно загружены прокси для {len(proxies)} стран")
            return proxy_manager

        except Exception as e:
            self.logger.error(f"Ошибка загрузки прокси: {str(e)}")
            raise

    def _load_main_queue(self) -> List[WorkUnit]:
        self.logger.info("Загрузка основной очереди из Excel файла")

        row = 1
        while True:
            link = self.sheet.cell(row=row, column=3).value
            if link is None:
                break

            link = link if link.startswith("http") else f"https://{link}"

            is_downloaded = bool(self.sheet.cell(row=row, column=16).value)
            if is_downloaded:
                row += 1
                continue

            title = self.sheet.cell(row=row, column=4).value
            lang = str(self.sheet.cell(row=row, column=5).value).lower()
            image_url = self.sheet.cell(row=row, column=8).value
            description = self.sheet.cell(row=row, column=9).value

            self.main_queue.append(
                WorkUnit(
                    link=link, title=title, lang=lang,
                    image_url=image_url, description=description,
                    is_downloaded=is_downloaded
                )
            )

            row += 1

        self.logger.info(f"Загружено {(row - 1)} задач в основную очередь")
        return self.main_queue

    def check_site_availability(self, link: str, proxy: bool = False, timeout: float = 5.0) -> bool:
        """
        Проверяет доступность веб-сайта, с возможностью использования прокси-сервера.
        """
        self.logger.debug(f"Проверка доступности сайта: {urlparse(link).netloc} (прокси: {'да' if proxy else 'нет'})")

        proxy_config = {
            "http": "socks5://127.0.0.1:2080",
            "https": "socks5://127.0.0.1:2080"
        } if proxy else None

        try:
            response = requests.get(
                link,
                timeout=timeout,
                proxies=proxy_config,
                allow_redirects=True
            )

            is_available = response.ok
            self.logger.debug(f"Сайт {urlparse(link).netloc} доступен: {'да' if is_available else 'нет'}")
            return is_available

        except requests.exceptions.RequestException as e:
            self.logger.warning(f"Ошибка проверки доступности сайта {urlparse(link).netloc}: {str(e)}")
            return False

    @staticmethod
    def download_image(image_url, save_path):
        """
        Скачивает изображение по URL и сохраняет в указанный путь.
        """
        logger = logging.getLogger(__name__)

        try:
            if not image_url:
                logger.warning("Ошибка: URL изображения не указан.")
                return False

            logger.info(f"Загрузка изображения из {urlparse(image_url).netloc} в {save_path}")

            response = requests.get(image_url, stream=True, timeout=10)
            response.raise_for_status()

            if 'image' not in response.headers.get('content-type', '').lower():
                logger.warning("Ошибка: По указанному URL нет изображения.")
                return False

            os.makedirs(os.path.dirname(save_path), exist_ok=True)

            with open(save_path, 'wb') as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)

            logger.info(f"Изображение успешно сохранено: {save_path}")
            return True

        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка при загрузке изображения: {e}")
        except IOError as e:
            logger.error(f"Ошибка при сохранении файла: {e}")
        except Exception as e:
            logger.error(f"Неизвестная ошибка: {e}")

        return False

    def _process_without_proxy(
        self,
        unit: Union[WorkUnit, WaitWorkUnit]
    ) -> bool:
        """
        Обрабатывает веб-сайт без использования прокси-сервера.
        """
        work = unit.work if isinstance(unit, WaitWorkUnit) else unit

        self.logger.info(f"Обработка без прокси: {work.link}")

        link_is_availably = self.check_site_availability(link=work.link)
        if not link_is_availably:
            self.logger.warning(f"Ссылка недоступна: {urlparse(work.link).netloc}")
            self.data.append(ResultWorkUnit(
                status="error", unit=work,
                timestamp=datetime.now(),
                context="Ссылка недоступна"
            ))
            return False

        if isinstance(unit, WaitWorkUnit):
            return self._process_browser(unit=unit)
        else:
            return self._process_browser(unit=work)

    def _process_with_proxy(
        self,
        unit: Union[WorkUnit, WaitWorkUnit],
        proxy: Optional[ProxyUnit] = None
    ) -> bool:
        """
        Обрабатывает веб-сайт через прокси-сервер с проверкой доступности и последующей обработкой в браузере.
        """
        if isinstance(unit, WaitWorkUnit):
            work = unit.work
            proxy = unit.proxy
        else:
            work = unit

        self.logger.info(f"Обработка с прокси {work.lang if proxy else 'None'}: {urlparse(work.link).netloc}")

        try:
            with Proxy(
                host=proxy.host, port=proxy.port,
                username=proxy.username, password=proxy.password
            ):
                link_is_availably = self.check_site_availability(link=work.link, proxy=True)
                if not link_is_availably:
                    self.logger.warning(f"Ссылка недоступна через прокси: {urlparse(work.link).netloc}")
                    self.data.append(ResultWorkUnit(
                        status="error", unit=work,
                        timestamp=datetime.now(),
                        context="Ссылка недоступна"
                    ))
                    return False

                if isinstance(unit, WaitWorkUnit):
                    return self._process_browser(unit=unit)
                else:
                    return self._process_browser(unit=work, proxy=proxy)

        except Exception as e:
            self.logger.error(f"Ошибка обработки с прокси: {str(e)}")
            return False

    def _process_browser(
        self,
        unit: Union[WaitWorkUnit, WorkUnit],
        proxy: Optional[ProxyUnit] = None
    ) -> bool:
        """
        Обрабатывает веб-сайт с использованием браузера, включая проверки и сохранение результатов.
        """
        if isinstance(unit, WaitWorkUnit):
            work = unit.work
            proxy = unit.proxy
        else:
            work = unit

        self.logger.info(f"Обработка в браузере: {urlparse(work.link).netloc} (язык: {work.lang}, прокси: {work.lang if proxy else 'нет'})")

        try:
            if work.lang != "ru":
                with Browser(proxy=None, headless=True) as browser:
                    browser.goto(work.link, delay=5)
                    main_title = browser.page.title()
                    self.logger.debug(f"Основной заголовок для {urlparse(work.link).netloc}: {main_title}")

            mobile_browser = (
                MobileBrowser(
                    locale=proxy.locale, longitude=proxy.longitude,
                    lantitude=proxy.lantitude, zipcode=proxy.zipcode
                )
                if proxy else
                MobileBrowser(proxy=None)
            )

            with mobile_browser as browser:
                browser.goto(work.link, delay=5)

                if work.lang != "ru":
                    secret_title = browser.page.title()
                    self.logger.debug(f"Скрытый заголовок для {urlparse(work.link).netloc}: {secret_title}")

                    if main_title == secret_title:
                        if isinstance(unit, WaitWorkUnit):
                            if (unit.attempts - 1) == 0:
                                self.logger.warning(f"Ссылка устарела после всех попыток: {urlparse(work.link).netloc}")
                                self.data.append(ResultWorkUnit(
                                    status="error", unit=work,
                                    timestamp=datetime.now(),
                                    context="Ссылка устарела"
                                ))

                            if (unit.attempts - 1) > 0:
                                unit.attempts = unit.attempts - 1
                                unit.timestamp=(datetime.now() + timedelta(minutes=7))
                                self.side_queue.append(unit)
                                self.logger.info(f"Добавлено в очередь повторных попыток (осталось попыток: {unit.attempts}): {urlparse(work.link).netloc}")

                            return False
                        else:
                            self.side_queue.append(WaitWorkUnit(
                                work=work, proxy=proxy, attempts=3,
                                timestamp=datetime.now() + timedelta(minutes=7)
                            ))
                            self.logger.info(f"Добавлено в очередь повторных попыток: {urlparse(work.link).netloc}")

                            return False

                if not browser.download_website("website"):
                    DirManager.clear_directory(Path("temp"))
                    self.logger.error(f"Ошибка загрузки сайта: {urlparse(work.link).netloc}")
                    self.data.append(ResultWorkUnit(
                        status="error", unit=work, timestamp=datetime.now(),
                        context="Проблемы при загрузке сайта",
                    ))
                    return False

                # try:
                #     browser.screenshot("temp/screenshot.png")
                #     self.logger.debug("Скриншот сохранен в PNG")
                # except:
                # browser.my_screenshot("temp/screenshot_1.pdf")

                browser.pdf("temp/screenshot.pdf")
                self.logger.debug(f"Скриншот сохранен в PDF")

            self.download_image(work.image_url, "temp/image.png")
            self._save_info_file(work.link, work.title, work.description)
            website_path = DirManager.move_to_numbered_dir(Path("temp"), Path(f"websites/{work.lang}"))

            self.logger.info(f"Успешно обработано: {urlparse(work.link).netloc}. Сохранено в: {website_path}")
            self.data.append(ResultWorkUnit(
                status="ok", path=website_path,
                unit=work, timestamp=datetime.now()
            ))

            return True

        except Exception as e:
            self.logger.error(f"Ошибка обработки в браузере для {urlparse(work.link).netloc}: {str(e)}")
            if isinstance(unit, WaitWorkUnit):
                if (unit.attempts - 1) == 0:
                    self.logger.warning(f"Ссылка устарела после всех попыток: {urlparse(work.link).netloc}")
                    self.data.append(ResultWorkUnit(
                        status="error", unit=work,
                        timestamp=datetime.now(),
                        context="Ссылка устарела"
                    ))

                if (unit.attempts - 1) > 0:
                    unit.attempts = unit.attempts - 1
                    unit.timestamp=(datetime.now() + timedelta(minutes=7))
                    self.side_queue.append(unit)
                    self.logger.info(f"Добавлено в очередь повторных попыток (осталось попыток: {unit.attempts}): {urlparse(work.link).netloc}")

            else:
                self.side_queue.append(WaitWorkUnit(
                    work=work, proxy=proxy, attempts=3,
                    timestamp=datetime.now() + timedelta(minutes=7)
                ))
                self.logger.info(f"Добавлено в очередь повторных попыток: {urlparse(work.link).netloc}")

            return False

    def _save_info_file(self, link: str, title: str, description: str) -> None:
        """Создает файл с информацией о сайте."""
        self.logger.debug(f"Сохранение info файла для {urlparse(link).netloc}")
        data = {
            "url": link,
            "description": description if description != "null" else title,
            "time": datetime.now().isoformat()
        }
        FileManager.write_file(data, "temp/info.txt")

    def process_all(self):
        """
        Основной Основной метод для обработки всех задач из основной и отложенной очередей.
        """
        self.logger.info("Начало обработки всех задач")
        start_time = datetime.now()
        processed_count = 0
        error_count = 0

        queue = self.main_queue[:]

        while len(queue) != 0:
            work = queue.pop(0)
            self.logger.info("==============================================")
            self.logger.info(f"Обработка задачи: {urlparse(work.link).netloc} (язык: {work.lang})")

            try:
                if work.lang != "ru":
                    if work.lang not in self.proxy_manager.regions:
                        self.logger.warning(f"Нет прокси для языка: {work.lang}")
                        self.data.append(ResultWorkUnit(
                            status="error", unit=work,
                            timestamp=datetime.now(),
                            context="Нет прокси для обработки ссылки"
                        ))
                        error_count += 1
                        continue

                    proxy = self.proxy_manager.get_proxy(country_name=work.lang)
                    if self._process_with_proxy(unit=work, proxy=proxy):
                        processed_count += 1
                    else:
                        error_count += 1

                else:
                    if self._process_without_proxy(unit=work):
                        processed_count += 1
                    else:
                        error_count += 1

                current_time = datetime.now()
                for unit in list(self.side_queue):
                    if unit.timestamp <= current_time:
                        self.side_queue.remove(unit)
                        self.logger.info("==============================================")
                        self.logger.info(f"Обработка задачи из очереди повторных попыток: {urlparse(unit.work.link).netloc}")

                        if unit.proxy:
                            if self._process_with_proxy(unit=unit):
                                processed_count += 1
                            else:
                                error_count += 1
                        else:
                            if self._process_without_proxy(unit=unit):
                                processed_count += 1
                            else:
                                error_count += 1

            except Exception as e:
                self.logger.error(f"Ошибка обработки задачи {urlparse(work.link).netloc}: {str(e)}")
                error_count += 1

        while self.side_queue:
            current_time = datetime.now()
            for unit in list(self.side_queue):
                if unit.timestamp <= current_time:
                    self.side_queue.remove(unit)
                    self.logger.info("==============================================")
                    self.logger.info(f"Обработка оставшихся задач из очереди повторных попыток: {urlparse(unit.work.link).netloc}")

                    if unit.proxy:
                        if self._process_with_proxy(unit=unit):
                            processed_count += 1
                        else:
                            error_count += 1
                    else:
                        if self._process_without_proxy(unit=unit):
                            processed_count += 1
                        else:
                            error_count += 1

            if self.side_queue:
                sleep_time = (self.side_queue[0].timestamp - datetime.now()).total_seconds()
                if sleep_time > 0:
                    self.logger.info(f"Ожидание {sleep_time} секунд перед следующей задачей...")
                    time.sleep(sleep_time)
            else:
                continue

        duration = (datetime.now() - start_time).total_seconds()
        self.logger.info(f"Обработка завершена. Успешно: {processed_count}, Ошибок: {error_count}, Время: {duration:.2f} секунд")

    def create_excel(self, save_path: str) -> bool:
        """
        Создает Excel-отчет с результатами обработки сайтов и сохраняет его по указанному пути.

        Args:
            save_path (str): Путь для сохранения Excel-файла

        Returns:
            bool: True если отчет успешно создан, False в случае ошибки
        """
        self.logger.info(f"Начало создания Excel отчета. Путь сохранения: {save_path}")

        # Проверка наличия данных
        if not self.data:
            self.logger.warning("Прерывание создания отчета: отсутствуют данные для отчета")
            return False

        # Преобразование пути и проверка директории
        try:
            save_path = Path(save_path).absolute()
            parent_dir = save_path.parent

            if not parent_dir.exists():
                self.logger.info(f"Создание директории для отчета: {parent_dir}")
                parent_dir.mkdir(parents=True, exist_ok=True)

            if not parent_dir.is_dir():
                self.logger.error(f"Указанный путь не является директорией: {parent_dir}")
                return False
        except Exception as e:
            self.logger.error(f"Ошибка обработки пути сохранения: {str(e)}")
            return False

        # Создание и заполнение Excel-файла
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты обработки"

            # Заголовки столбцов
            headers = [
                "Статус", "Контекст", "Ссылка",
                "Язык", "Путь", "Загружено", "Время"
            ]
            ws.append(headers)

            # Форматирование заголовков
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Заполнение данными
            for idx, unit in enumerate(self.data, start=2):
                try:
                    row_data = [
                        str(unit.status),
                        str(unit.context),
                        str(unit.unit.link),
                        str(unit.unit.lang),
                        str(unit.path) if unit.path else "",
                        "Да" if unit.path else "Нет",
                        unit.timestamp.isoformat() if hasattr(unit, 'timestamp') else ""
                    ]
                    ws.append(row_data)

                    # Автоподбор ширины столбцов
                    for col in range(1, len(row_data) + 1):
                        cell = ws.cell(row=idx, column=col)
                        if len(str(cell.value)) > ws.column_dimensions[get_column_letter(col)].width:
                            ws.column_dimensions[get_column_letter(col)].width = len(str(cell.value)) + 2

                except Exception as unit_error:
                    self.logger.warning(f"Ошибка обработки единицы данных {idx}: {str(unit_error)}")
                    continue

            # Сохранение файла
            wb.save(save_path)
            self.logger.info(f"Excel отчет успешно создан. Размер: {save_path.stat().st_size/1024:.2f} KB")
            return True

        except PermissionError:
            self.logger.error(f"Ошибка доступа: невозможно сохранить файл {save_path}")
            return False
        except Exception as e:
            self.logger.error(f"Критическая ошибка при создании отчета: {str(e)}", exc_info=True)
            return False


if __name__ == "__main__":
    # Настройка парсера аргументов командной строки
    parser = argparse.ArgumentParser(description='Обработка данных из Excel файла')
    parser.add_argument('--excel_path', type=str, required=True,
                        help='Путь к Excel файлу с исходными данными')
    parser.add_argument('--result_path', type=str, required=True,
                        help='Путь для сохранения результата (Excel файл)')

    args = parser.parse_args()

    try:
        processor = WebsiteProcessor(excel_path=args.excel_path)
        processor.process_all()
    except Exception as e:
        logging.getLogger(__name__).error(f"Критическая ошибка при выполнении: {str(e)}", exc_info=True)
    finally:
        processor.create_excel(save_path=args.result_path)
